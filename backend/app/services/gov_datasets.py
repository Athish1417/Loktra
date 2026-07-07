"""Real government-dataset importer (Phase 4).

Reads official CSV / XLSX / XLS files dropped into the datasets/ folders, detects
columns robustly (headers vary between sources), normalises the useful fields, and
stores them in the per-type record tables. Invalid rows are skipped, never fatal —
if a real dataset is missing or partial the app keeps working on sample data.
"""
import csv
import io
import re
import shutil
from pathlib import Path
from typing import IO, Dict, List, Optional, Tuple

from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.dataset import DatasetMeta
from app.models.gov_dataset import (
    DatasetSource,
    ElectionConstituencyRecord,
    LgdLocationRecord,
    NfhsRecord,
    PincodeRecord,
    PopulationRecord,
)

SOURCE_TYPES = ["census", "lgd", "pincode", "nfhs", "election", "imports"]
_ALLOWED_EXT = {".csv", ".xlsx", ".xls"}

# Records table per source type — used to delete a source's rows on re-import/delete.
_RECORD_MODELS = {
    "census": PopulationRecord,
    "imports": PopulationRecord,
    "lgd": LgdLocationRecord,
    "pincode": PincodeRecord,
    "nfhs": NfhsRecord,
    "election": ElectionConstituencyRecord,
}


class DatasetImportError(ValueError):
    """Raised for unreadable files or unimportable content."""


# --------------------------------------------------------------------------- #
# Column detection (robust to messy government headers)
# --------------------------------------------------------------------------- #
FIELD_ALIASES: Dict[str, List[str]] = {
    "state": ["state", "statename", "stateut", "stname", "stnm"],
    "district": ["district", "districtname", "dtname", "dist", "distname", "dtnm"],
    "subdistrict": ["subdistrict", "subdist", "taluk", "taluka", "tehsil", "block", "mandal"],
    "village": ["village", "villagename", "town", "townvillage", "locality", "villagetown"],
    "constituency": [
        "constituency", "constituencyname", "pcname", "acname", "pc", "ac",
        "parliamentaryconstituency", "assemblyconstituency", "pcno", "acno",
    ],
    "population": ["population", "totalpopulation", "totpop", "totp", "persons", "totalpersons"],
    "households": ["households", "numberofhouseholds", "totalhouseholds", "noofhouseholds", "hh"],
    "area": ["area", "areasqkm", "totalarea", "geographicalarea", "areainhectares"],
    "pincode": ["pincode", "pin", "pincode6digit", "pinno"],
    "officename": ["officename", "postoffice", "office", "poname"],
}


def _norm(text) -> str:
    return re.sub(r"[^a-z0-9]", "", str(text or "").lower())


def detect_columns(headers: List[str], fields: List[str]) -> Dict[str, int]:
    """Map each wanted field to a column index (exact alias, then substring)."""
    norm_headers = [_norm(h) for h in headers]
    mapping: Dict[str, int] = {}
    for field in fields:
        aliases = FIELD_ALIASES.get(field, [field])
        idx = next((i for i, h in enumerate(norm_headers) if h in aliases), None)
        if idx is None:  # looser: alias appears inside the header
            idx = next(
                (i for i, h in enumerate(norm_headers)
                 if any(a in h for a in aliases)),
                None,
            )
        if idx is not None:
            mapping[field] = idx
    return mapping


def _to_int(value) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).replace(",", "").strip()))
    except (TypeError, ValueError):
        return None


def _to_float(value) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _cell(row, mapping, field):
    idx = mapping.get(field)
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    return val.strip() if isinstance(val, str) else val


# --------------------------------------------------------------------------- #
# Folder scanning + file reading
# --------------------------------------------------------------------------- #
def _base_dir() -> Path:
    return Path(settings.DATASETS_DIR).resolve()


def _safe_path(rel_path: str) -> Path:
    """Resolve a path and confirm it stays inside the datasets directory."""
    base = _base_dir()
    target = (base / rel_path).resolve()
    if base not in target.parents and target != base:
        raise DatasetImportError("Path is outside the datasets directory.")
    if not target.is_file():
        raise DatasetImportError(f"File not found: {rel_path}")
    return target


def scan() -> List[dict]:
    """List every importable file under the dataset folders."""
    base = _base_dir()
    found = []
    for source_type in SOURCE_TYPES:
        folder = base / source_type
        if not folder.is_dir():
            continue
        for f in sorted(folder.iterdir()):
            if f.is_file() and f.suffix.lower() in _ALLOWED_EXT:
                found.append({
                    "source_type": source_type,
                    "file_name": f.name,
                    "path": f"{source_type}/{f.name}",
                    "ext": f.suffix.lower(),
                    "size_kb": round(f.stat().st_size / 1024, 1),
                })
    return found


def save_upload(fileobj: IO[bytes], filename: str, source_type: str = "imports") -> str:
    """Save an uploaded file into datasets/<source_type>/ and return its rel path.

    Additive plumbing so the browser can upload a file the existing folder-based
    importer (preview / import_file) then handles by path — the importer itself is
    unchanged.
    """
    st = source_type if source_type in SOURCE_TYPES else "imports"
    folder = _base_dir() / st
    folder.mkdir(parents=True, exist_ok=True)
    safe = re.sub(r"[^A-Za-z0-9._-]", "_", Path(filename or "upload").name) or "upload.csv"
    with open(folder / safe, "wb") as out:
        shutil.copyfileobj(fileobj, out)
    return f"{st}/{safe}"


_DETECT_FIELDS = [
    "state", "district", "subdistrict", "village", "population",
    "households", "area", "pincode", "officename", "constituency",
]


# Filename tokens (normalised, separators stripped) that name an official dataset.
# Official government files are reliably named, so the filename is the strongest signal.
_NAME_SIGNATURES = [
    ("nfhs", "nfhs"),
    ("pincode", "pincode"),
    ("postoffice", "pincode"),
    ("postal", "pincode"),
    ("lgd", "lgd"),
    ("census", "census"),
    ("election", "election"),
    ("constituency", "election"),
]

# Header phrases unique to an NFHS factsheet. These are long and specific, so a
# substring match here is safe (unlike the short "pin"/"ac" field aliases).
_NFHS_HEADER_HINTS = [
    "householdssurveyed", "womenage", "menage", "sexratio", "antenatal", "nfhs",
]


def _has_exact_field(norm_headers: List[str], field: str) -> bool:
    """True only if a whole header cell equals one of the field's aliases.

    Uses exact-alias matching (no loose substring) so unrelated columns like
    "...check-up in the..." no longer look like a PIN code via the short "pin" alias.
    """
    aliases = FIELD_ALIASES.get(field, [field])
    return any(h in aliases for h in norm_headers)


def detect_source_type(rel_path: str) -> str:
    """Best-guess the dataset type from the filename, then column signatures.

    Filename hints win first; otherwise strict header signatures decide. PIN code is
    only chosen when a header is genuinely a pincode/post-office field (or the file is
    named so), never from a short alias appearing inside an unrelated indicator name.
    """
    path = _safe_path(rel_path)
    name = _norm(path.name)
    for token, source_type in _NAME_SIGNATURES:
        if token in name:
            return source_type

    headers, _, _, _ = read_table(path, _DETECT_FIELDS, max_rows=1)
    norm_headers = [_norm(h) for h in headers]

    if any(hint in h for h in norm_headers for hint in _NFHS_HEADER_HINTS):
        return "nfhs"
    if _has_exact_field(norm_headers, "pincode") or _has_exact_field(norm_headers, "officename"):
        return "pincode"
    if _has_exact_field(norm_headers, "constituency"):
        return "election"
    if _has_exact_field(norm_headers, "population"):
        return "census"
    if _has_exact_field(norm_headers, "village"):
        return "lgd"
    return "imports"


def _is_blank(row) -> bool:
    return all(str(c).strip() == "" for c in row)


def _read_sheets(path: Path) -> List[Tuple[str, List[list]]]:
    """Return [(sheet_name, rows)] for a CSV/XLSX/XLS file (all sheets)."""
    ext = path.suffix.lower()
    if ext == ".csv":
        text = path.read_bytes().decode("utf-8-sig", errors="replace")
        return [("csv", list(csv.reader(io.StringIO(text))))]
    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:  # pragma: no cover
            raise DatasetImportError("openpyxl is required to read .xlsx files.")
        wb = load_workbook(path, read_only=True, data_only=True)
        out = [
            (ws.title, [["" if c is None else c for c in row]
                        for row in ws.iter_rows(values_only=True)])
            for ws in wb.worksheets
        ]
        wb.close()
        return out
    if ext == ".xls":
        try:
            import xlrd
        except ImportError:  # pragma: no cover
            raise DatasetImportError("xlrd is required to read legacy .xls files.")
        book = xlrd.open_workbook(str(path))
        return [
            (s.name, [s.row_values(r) for r in range(s.nrows)])
            for s in (book.sheet_by_index(i) for i in range(book.nsheets))
        ]
    raise DatasetImportError(f"Unsupported file type: {ext}")


def _detect_header_row(rows: List[list], fields: List[str], scan: int = 25) -> int:
    """Find the header row, skipping title/note rows government files often prepend.

    Picks the row (within the first `scan`) whose cells match the most target
    fields; if nothing matches, the non-blank row with the most filled cells.
    """
    limit = min(len(rows), scan)
    best_idx, best_score = 0, -1
    for i in range(limit):
        if _is_blank(rows[i]):
            continue
        score = len(detect_columns([str(c) for c in rows[i]], fields))
        if score > best_score:
            best_idx, best_score = i, score
    if best_score <= 0 and limit:
        best_idx = max(
            range(limit),
            key=lambda i: sum(1 for c in rows[i] if str(c).strip() != ""),
        )
    return best_idx


def read_table(
    path: Path, fields: Optional[List[str]] = None, max_rows: Optional[int] = None
) -> Tuple[List[str], List[list], str, List[str]]:
    """Return (headers, data_rows, chosen_sheet, all_sheet_names).

    Chooses the sheet whose detected header matches the most target fields,
    detects the header row past any title/notes, and drops blank rows.
    """
    sheets = _read_sheets(path)
    fields = fields or []
    sheet_names = [name for name, _ in sheets]

    def sheet_score(rows: List[list]) -> Tuple[int, int]:
        if not rows:
            return (-1, 0)
        hi = _detect_header_row(rows, fields)
        matched = len(detect_columns([str(c) for c in rows[hi]], fields))
        return (matched, len(rows))

    name, rows = max(sheets, key=lambda s: sheet_score(s[1])) if sheets else ("", [])
    if not rows:
        return [], [], name, sheet_names

    hi = _detect_header_row(rows, fields)
    headers = [str(h).strip() for h in rows[hi]]
    data = [r for r in rows[hi + 1:] if not _is_blank(r)]
    if max_rows is not None:
        data = data[:max_rows]
    return headers, data, name, sheet_names


# --------------------------------------------------------------------------- #
# Preview
# --------------------------------------------------------------------------- #
_FIELDS_BY_TYPE = {
    "census": ["state", "district", "subdistrict", "village", "population", "households", "area"],
    "lgd": ["state", "district", "subdistrict", "village"],
    "pincode": ["pincode", "officename", "district", "state"],
    "nfhs": ["state", "district"],
    "election": ["state", "district", "constituency"],
    "imports": ["state", "district", "subdistrict", "village", "population", "households", "area"],
}


def _fields_for(source_type: str) -> List[str]:
    return _FIELDS_BY_TYPE.get(source_type, _FIELDS_BY_TYPE["imports"])


def preview(rel_path: str, source_type: str, limit: int = 15) -> dict:
    path = _safe_path(rel_path)
    fields = _fields_for(source_type)
    headers, rows, sheet, sheet_names = read_table(path, fields, max_rows=limit)
    mapping = detect_columns(headers, fields)
    return {
        "file_name": path.name,
        "source_type": source_type,
        "sheet": sheet,
        "sheet_names": sheet_names,
        "columns": headers,
        "detected_fields": {f: headers[i] for f, i in mapping.items()},
        "sample_rows": [list(r) for r in rows],
    }


# --------------------------------------------------------------------------- #
# Import
# --------------------------------------------------------------------------- #
def import_file(
    db: Session,
    *,
    rel_path: str,
    source_type: str,
    is_official: bool = True,
    dataset_name: Optional[str] = None,
    source_name: Optional[str] = None,
    source_department: Optional[str] = None,
    source_url: Optional[str] = None,
    replace: bool = False,
) -> dict:
    if source_type not in SOURCE_TYPES:
        raise DatasetImportError(f"Unknown source_type '{source_type}'.")
    path = _safe_path(rel_path)

    # Avoid duplicate imports of the same file unless the caller asks to replace.
    existing = (
        db.query(DatasetSource)
        .filter(DatasetSource.file_name == path.name,
                DatasetSource.source_type == source_type)
        .first()
    )
    if existing and not replace:
        return {
            "source_type": source_type, "file_name": path.name,
            "imported": 0, "skipped": 0, "record_count": existing.record_count,
            "status": "skipped_duplicate", "mode": current_mode(db),
            "detail": "This file is already imported. Pass replace=true to re-import.",
        }
    if existing and replace:
        _delete_source(db, existing)

    fields = _fields_for(source_type)
    headers, rows, sheet, sheet_names = read_table(path, fields)
    if not headers:
        raise DatasetImportError("File has no readable header row.")
    mapping = detect_columns(headers, fields)

    source = DatasetSource(
        dataset_name=dataset_name or f"{source_type.upper()} dataset",
        source_name=source_name or path.name,
        source_department=source_department,
        source_type=source_type,
        file_name=path.name,
        file_path=f"{source_type}/{path.name}",
        source_url=source_url,
        is_official=is_official,
        record_count=0,
        import_status="success",
    )
    db.add(source)
    db.flush()  # get source.id

    builder = _BUILDERS.get(source_type, _build_population)
    imported, skipped = 0, 0
    for row in rows:
        try:
            rec = builder(row, headers, mapping, source.id)
        except Exception:
            rec = None
        if rec is None:
            skipped += 1
            continue
        db.add(rec)
        imported += 1

    source.record_count = imported
    source.import_status = (
        "success" if imported and not skipped
        else "partial" if imported
        else "failed"
    )

    # Flip the dataset mode to official only when a real file actually imported rows.
    if is_official and imported > 0:
        _mark_official(db, source)

    db.commit()
    return {
        "source_id": source.id,
        "source_type": source_type,
        "file_name": path.name,
        "sheet": sheet,
        "sheet_names": sheet_names,
        "columns": headers,
        "detected_fields": {f: headers[i] for f, i in mapping.items()},
        "imported": imported,
        "skipped": skipped,
        "record_count": imported,
        "import_status": source.import_status,
        "status": "imported",
        "mode": current_mode(db),
    }


def _delete_source_records(db: Session, source: DatasetSource) -> None:
    model = _RECORD_MODELS.get(source.source_type)
    if model is not None:
        db.query(model).filter(model.source_id == source.id).delete(
            synchronize_session=False
        )


def _delete_source(db: Session, source: DatasetSource) -> None:
    _delete_source_records(db, source)
    db.delete(source)
    db.flush()


def delete_source(db: Session, source_id: int) -> bool:
    """Delete an imported dataset source + its records. Re-labels mode if needed."""
    source = db.query(DatasetSource).filter(DatasetSource.id == source_id).first()
    if not source:
        return False
    _delete_source(db, source)
    # If no official data remains, honestly drop back to sample mode.
    if not has_official_data(db):
        _reset_to_sample(db)
    db.commit()
    return True


def _build_population(row, headers, mapping, source_id):
    state = _cell(row, mapping, "state")
    district = _cell(row, mapping, "district")
    if not state and not district:
        return None  # nothing to anchor on
    return PopulationRecord(
        source_id=source_id,
        state=state, district=district,
        subdistrict=_cell(row, mapping, "subdistrict"),
        village=_cell(row, mapping, "village"),
        population=_to_int(_cell(row, mapping, "population")),
        households=_to_int(_cell(row, mapping, "households")),
        area=_to_float(_cell(row, mapping, "area")),
    )


def _build_lgd(row, headers, mapping, source_id):
    state = _cell(row, mapping, "state")
    district = _cell(row, mapping, "district")
    village = _cell(row, mapping, "village")
    if not any([state, district, village]):
        return None
    return LgdLocationRecord(
        source_id=source_id, state=state, district=district,
        subdistrict=_cell(row, mapping, "subdistrict"), village=village,
    )


def _build_pincode(row, headers, mapping, source_id):
    pincode = _cell(row, mapping, "pincode")
    if not pincode:
        return None
    return PincodeRecord(
        source_id=source_id, pincode=str(pincode).strip(),
        office_name=_cell(row, mapping, "officename"),
        district=_cell(row, mapping, "district"),
        state=_cell(row, mapping, "state"),
    )


def _build_nfhs(row, headers, mapping, source_id):
    state = _cell(row, mapping, "state")
    district = _cell(row, mapping, "district")
    if not state and not district:
        return None
    skip = {mapping.get("state"), mapping.get("district")}
    indicators = {
        headers[i]: row[i]
        for i in range(min(len(headers), len(row)))
        if i not in skip and str(row[i]).strip() != ""
    }
    return NfhsRecord(
        source_id=source_id,
        level="district" if district else "state",
        state=state, district=district, indicators=indicators,
    )


def _build_election(row, headers, mapping, source_id):
    state = _cell(row, mapping, "state")
    district = _cell(row, mapping, "district")
    constituency = _cell(row, mapping, "constituency")
    if not any([state, district, constituency]):
        return None
    skip = {mapping.get("state"), mapping.get("district"), mapping.get("constituency")}
    extra = {
        headers[i]: row[i]
        for i in range(min(len(headers), len(row)))
        if i not in skip and str(row[i]).strip() != ""
    }
    return ElectionConstituencyRecord(
        source_id=source_id, state=state, district=district,
        constituency=constituency, extra=extra,
    )


_BUILDERS = {
    "census": _build_population,
    "imports": _build_population,
    "lgd": _build_lgd,
    "pincode": _build_pincode,
    "nfhs": _build_nfhs,
    "election": _build_election,
}


# --------------------------------------------------------------------------- #
# Dataset mode + provenance
# --------------------------------------------------------------------------- #
def _get_meta(db: Session) -> DatasetMeta:
    meta = db.query(DatasetMeta).first()
    if not meta:
        meta = DatasetMeta()
        db.add(meta)
        db.flush()
    return meta


def _mark_official(db: Session, source: DatasetSource) -> None:
    import datetime

    meta = _get_meta(db)
    meta.dataset_name = source.dataset_name
    meta.source_name = source.source_name
    meta.is_official = True
    meta.imported_at = datetime.datetime.utcnow()
    db.add(meta)


def _reset_to_sample(db: Session) -> None:
    import datetime

    meta = _get_meta(db)
    meta.dataset_name = "Sample Government-Style Dataset"
    meta.source_name = "Sample (demo) data — not an official source"
    meta.is_official = False
    meta.imported_at = datetime.datetime.utcnow()
    db.add(meta)


def has_official_data(db: Session) -> bool:
    return (
        db.query(DatasetSource)
        .filter(DatasetSource.is_official.is_(True), DatasetSource.record_count > 0)
        .first()
        is not None
    )


def current_mode(db: Session) -> str:
    return (
        "Official Government Dataset Imported"
        if has_official_data(db)
        else "Sample Government-Style Dataset"
    )


def list_sources(db: Session) -> List[DatasetSource]:
    return db.query(DatasetSource).order_by(DatasetSource.imported_at.desc()).all()


# --------------------------------------------------------------------------- #
# Location intelligence — used by the AI priority engine (safe, optional)
# --------------------------------------------------------------------------- #
def real_population_for(
    db: Session, state: Optional[str], district: Optional[str]
) -> Optional[int]:
    """Real population for a district from imported Census data, or None.

    Sums matching PopulationRecord rows (census may be village/district level).
    Uses fuzzy case-insensitive name matching; returns None when nothing matches
    so callers fall back to the sample dataset.
    """
    if not district:
        return None
    q = db.query(PopulationRecord).filter(PopulationRecord.population.isnot(None))
    q = q.filter(PopulationRecord.district.ilike(district.strip()))
    if state:
        q = q.filter(PopulationRecord.state.ilike(state.strip()))
    rows = q.all()
    if not rows:
        return None
    total = sum(r.population for r in rows if r.population)
    return total or None


def context_for(
    db: Session,
    *,
    state: Optional[str] = None,
    district: Optional[str] = None,
    constituency: Optional[str] = None,
) -> dict:
    """Which imported official datasets have context for this location.

    Cheap existence checks used to enrich (and explain) the AI assessment. Always
    safe: returns all-False when nothing is imported.
    """
    ctx = {"census": False, "nfhs": False, "election": False}
    if district:
        ctx["census"] = (
            db.query(PopulationRecord.id)
            .filter(PopulationRecord.district.ilike(district.strip()))
            .first() is not None
        )
        ctx["nfhs"] = (
            db.query(NfhsRecord.id)
            .filter(NfhsRecord.district.ilike(district.strip()))
            .first() is not None
        )
    if constituency:
        ctx["election"] = (
            db.query(ElectionConstituencyRecord.id)
            .filter(ElectionConstituencyRecord.constituency.ilike(constituency.strip()))
            .first() is not None
        )
    return ctx
